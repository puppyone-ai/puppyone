"""Landing ingest orchestration (preview + claim).

See ``__init__.py`` for the architecture rationale (Option C).
"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import hmac
import json
import logging
import time
from io import BytesIO
from pathlib import PurePosixPath
from uuid import UUID, uuid4

from src.config import settings
from src.connectors.mcp_endpoint.schemas import McpAccessItem
from src.connectors.mcp_endpoint.service import McpEndpointService
from src.exceptions import AppException, ErrorCode
from src.infra.s3.service import S3Service
from src.platform.entitlements.service import EntitlementService
from src.platform.landing import tickets
from src.platform.landing.registry import ToolSpec, get_tool_spec
from src.platform.organization.dependencies import resolve_org_id
from src.platform.project.control_plane import ProjectControlPlaneService
from src.platform.project.models import Project
from src.platform.project.orchestration import create_project_with_tree
from src.repo.access_credentials import AccessCredentialRepository
from src.repo.scope_service import ScopeService
from src.version_engine.write_engine.engine import VersionWriteEngine

logger = logging.getLogger(__name__)

# Temp S3 prefix for anonymous previews. Put an S3 lifecycle rule on this
# prefix (e.g. expire after 3 days) so abandoned previews are reaped for free.
LANDING_PREFIX = "landing"
PREVIEW_TTL_SECONDS = 60 * 60 * 24  # 24h — must be <= the S3 lifecycle window
PREVIEW_EXCERPT_CHARS = 1500
SRC_PRESIGN_SECONDS = 900  # presigned URL lifetime handed to the OCR provider


def _landing_mcp_api_key(*, ticket_id: str, user_id: str) -> str:
    """Derive the replay-stable one-time MCP secret for a signed claim.

    Only its HMAC hash is persisted by the credential repository.  The signed
    ticket and authenticated user deterministically recover the same high-
    entropy value after a successful response is lost, so idempotent replay
    never creates a second endpoint or silently returns an empty credential.
    """

    secret = settings.JWT_SECRET or ""
    if not secret:
        raise tickets.TicketError("JWT_SECRET is not configured")
    digest = hmac.new(
        secret.encode(),
        f"landing-mcp:v1:{ticket_id}:{user_id}".encode(),
        hashlib.sha256,
    ).digest()
    encoded = base64.urlsafe_b64encode(digest).rstrip(b"=").decode("ascii")
    return f"mcp_{encoded}"


def _safe_name(filename: str | None) -> str:
    # Defend against path traversal / odd names in the S3 key and repo path.
    base = PurePosixPath((filename or "file").replace("\\", "/")).name
    return base or "file"


_HTTP_METHODS = {"get", "post", "put", "patch", "delete", "head", "options"}


def _openapi_to_markdown(content: bytes) -> str:
    """Render an OpenAPI/Swagger spec (JSON or YAML) into a readable markdown
    summary the MCP fs tools can serve."""
    text = content.decode("utf-8", errors="replace")
    spec: object
    try:
        spec = json.loads(text)
    except Exception:
        import yaml

        spec = yaml.safe_load(text)
    if not isinstance(spec, dict):
        raise ValueError("Not a valid OpenAPI/Swagger document")

    info = spec.get("info") or {}
    lines = [f"# {info.get('title', 'API')} {info.get('version', '')}".strip()]
    if info.get("description"):
        lines += ["", str(info["description"])]

    servers = spec.get("servers") or []
    server_urls = [s.get("url", "") for s in servers if isinstance(s, dict)]
    if server_urls:
        lines += ["", "## Servers"] + [f"- {u}" for u in server_urls if u]

    paths = spec.get("paths") or {}
    if isinstance(paths, dict) and paths:
        lines += ["", "## Endpoints"]
        for path, ops in paths.items():
            if not isinstance(ops, dict):
                continue
            for method, op in ops.items():
                if method.lower() not in _HTTP_METHODS:
                    continue
                op = op or {}
                summary = op.get("summary") or op.get("operationId") or ""
                entry = f"- `{method.upper()} {path}`"
                lines.append(f"{entry} — {summary}" if summary else entry)

    components = spec.get("components") or {}
    schemas = components.get("schemas") if isinstance(components, dict) else None
    schemas = schemas or spec.get("definitions") or {}
    if isinstance(schemas, dict) and schemas:
        lines += ["", "## Schemas", ", ".join(sorted(schemas.keys()))]

    return "\n".join(lines).rstrip() + "\n"


def _rows_to_md_table(rows: list[list[str]], truncated: bool, max_rows: int) -> list[str]:
    out: list[str] = []
    if rows:
        width = max(len(r) for r in rows)
        pad = lambda r: r + [""] * (width - len(r))  # noqa: E731
        out.append("| " + " | ".join(pad(rows[0])) + " |")
        out.append("| " + " | ".join(["---"] * width) + " |")
        for r in rows[1:]:
            out.append("| " + " | ".join(c.replace("|", "\\|") for c in pad(r)) + " |")
    if truncated:
        out.append(f"\n_(truncated at {max_rows} rows)_")
    return out


def _pdf_extract_text(content: bytes) -> str:
    """Extract the text layer of a PDF via PyMuPDF (no external service)."""
    import fitz  # PyMuPDF

    parts: list[str] = []
    with fitz.open(stream=content, filetype="pdf") as doc:
        for page in doc:
            parts.append(page.get_text())
    return "\n\n".join(p for p in parts if p).strip()


def _excel_to_markdown(content: bytes, max_rows: int = 200, max_cols: int = 30) -> str:
    """Render each worksheet of an .xlsx workbook as a markdown table. CSV/TSV
    uploads (detected by NOT being a zip/xlsx) go through the CSV path — we
    branch on the magic bytes so a binary xlsx is never fed to the CSV reader."""
    # .xlsx/.xlsm are zip archives ("PK\x03\x04"); real CSV/TSV is text.
    if content[:2] != b"PK":
        return _csv_to_markdown(content, max_rows, max_cols)

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"## {ws.title}")
        rows: list[list[str]] = []
        truncated = False
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_rows:
                truncated = True
                break
            rows.append(["" if c is None else str(c) for c in row[:max_cols]])
        out += _rows_to_md_table(rows, truncated, max_rows)
        out.append("")
    return "\n".join(out).rstrip() + "\n"


def _csv_to_markdown(content: bytes, max_rows: int, max_cols: int) -> str:
    import csv
    import io

    text = content.decode("utf-8", errors="replace")
    sample = text[:4096]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    rows: list[list[str]] = []
    truncated = False
    for r, row in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter)):
        if r >= max_rows:
            truncated = True
            break
        rows.append([("" if c is None else str(c)) for c in row[:max_cols]])
    return "\n".join(["## Sheet1", *_rows_to_md_table(rows, truncated, max_rows)]).rstrip() + "\n"


class LandingService:
    def __init__(
        self,
        s3_service: S3Service,
        control_plane: ProjectControlPlaneService,
        entitlements: EntitlementService,
        version_engine: VersionWriteEngine,
    ):
        self._s3 = s3_service
        self._control_plane = control_plane
        self._entitlements = entitlements
        self._version_engine = version_engine

    # ── preview: parse + stash, NO db rows ───────────────────────────
    async def preview(
        self,
        *,
        tool_kind: str,
        filename: str,
        content: bytes,
        content_type: str | None,
    ) -> dict:
        spec = get_tool_spec(tool_kind)
        ticket_id = uuid4().hex
        name = _safe_name(filename)

        src_key = f"{LANDING_PREFIX}/{ticket_id}/source/{name}"
        await self._s3.upload_file(
            src_key, content, content_type or "application/octet-stream"
        )

        markdown = await self._parse(
            spec, src_key=src_key, content=content, ticket_id=ticket_id
        )

        md_name = (PurePosixPath(name).stem or "document") + spec.output_ext
        md_key = f"{LANDING_PREFIX}/{ticket_id}/parsed/{md_name}"
        await self._s3.upload_file(
            md_key, markdown.encode("utf-8"), "text/markdown; charset=utf-8"
        )

        exp = int(time.time()) + PREVIEW_TTL_SECONDS
        ticket = tickets.sign_ticket(
            {
                "tid": ticket_id,
                "kind": spec.kind,
                "md_key": md_key,
                "md_name": md_name,
                "src_name": name,
                "exp": exp,
            }
        )
        return {
            "ticket": ticket,
            "preview": {
                "filename": name,
                "tool_kind": spec.kind,
                "content_chars": len(markdown),
                "excerpt": markdown[:PREVIEW_EXCERPT_CHARS],
                "suggested_tools": ["fs_ls", "fs_cat", "fs_grep", "fs_glob"],
            },
            "expires_at": exp,
        }

    async def _parse(
        self, spec: ToolSpec, *, src_key: str, content: bytes, ticket_id: str
    ) -> str:
        if spec.parser == "passthrough":
            return content.decode("utf-8", errors="replace")
        if spec.parser == "openapi":
            return _openapi_to_markdown(content)
        if spec.parser == "excel":
            return _excel_to_markdown(content)
        if spec.parser == "pdf":
            # Text-layer PDFs extract directly (no API key). Only fall back to
            # the OCR provider for scanned/image-only PDFs that yield no text.
            text = _pdf_extract_text(content)
            if len(text.strip()) >= 40:
                return text
            return await self._ocr_parse(src_key, ticket_id) or (
                text
                or "_(No extractable text — this looks like a scanned PDF; "
                "configure an OCR provider to read it.)_"
            )
        if spec.parser == "ocr_parse":
            return await self._ocr_parse(src_key, ticket_id)
        raise ValueError(
            f"Unsupported parser '{spec.parser}' for tool '{spec.kind}'"
        )

    async def _ocr_parse(self, src_key: str, ticket_id: str) -> str:
        """Run the configured OCR provider over the stashed source. Returns ""
        (not an exception) when OCR is unavailable, so callers can fall back."""
        try:
            from src.ingest.file.ocr.factory import get_ocr_provider

            presigned = await self._s3.generate_presigned_download_url(
                src_key, expires_in=SRC_PRESIGN_SECONDS
            )
            parsed = await get_ocr_provider().parse_document(
                presigned, data_id=ticket_id
            )
            return parsed.markdown_content or ""
        except Exception:  # noqa: BLE001 - provider missing/misconfigured → caller falls back
            logger.warning("OCR parse unavailable", exc_info=True)
            return ""

    # ── claim: born-owned create-chain for the logged-in user ────────
    async def claim(self, *, ticket: str, user_id: str) -> dict:
        body = tickets.verify_ticket(ticket)  # raises TicketError
        spec = get_tool_spec(body["kind"])
        md_bytes = await self._s3.download_file(body["md_key"])
        content_sha256 = hashlib.sha256(md_bytes).hexdigest()

        try:
            operation_uuid = UUID(hex=str(body["tid"]))
        except (KeyError, TypeError, ValueError) as exc:
            raise tickets.TicketError("ticket has an invalid operation id") from exc
        if operation_uuid.version != 4:
            raise tickets.TicketError("ticket operation id is not UUIDv4")
        operation_key = str(operation_uuid)
        api_key = _landing_mcp_api_key(ticket_id=operation_key, user_id=user_id)

        org_id = resolve_org_id(None, user_id)
        project_limit = await asyncio.to_thread(
            self._entitlements.enforced_limit_value,
            org_id,
            "projects.max",
        )
        stem = PurePosixPath(body["src_name"]).stem or "Document"

        # Shared born-owned create-chain (create Project row + init tree),
        # same as platform/project/router.create_project. The container is built
        # here because it is reused below for the content write.
        from src.version_engine.bootstrap.dependencies import (
            build_worker_version_engine_container,
        )
        container = build_worker_version_engine_container()
        endpoint_holder: dict[str, dict] = {}

        async def initialize(project: Project) -> None:
            project_id = str(project.id)
            # A single named leaf scope == the "separate repo" the user sees,
            # and is the narrowest scope for everything written under it.
            await asyncio.to_thread(
                ScopeService().create,
                project_id=project_id,
                name=spec.scope_path,
                path=spec.scope_path,
                exclude=[],
                mode="rw",
            )

            file_path = f"{spec.scope_path}/{body['md_name']}"
            await container.write_commands().bulk_write(
                project_id,
                {file_path: md_bytes},
                actor=f"landing:{user_id}",
                message=f"Import {body['src_name']} ({spec.kind} → MCP)",
                source_channel="upload",
            )

            endpoint_holder["endpoint"] = await asyncio.to_thread(
                McpEndpointService().create_endpoint,
                project_id=project_id,
                name=f"{spec.kind}-mcp",
                path=spec.scope_path,
                description=f"MCP for {body['src_name']}",
                accesses=[
                    McpAccessItem(
                        path=spec.scope_path,
                        json_path="",
                        readonly=spec.readonly,
                    )
                ],
                created_by=user_id,
                api_key=api_key,
            )

        publication = await create_project_with_tree(
            control_plane=self._control_plane,
            version_engine=self._version_engine,
            operation_key=operation_key,
            name=f"{stem}{spec.name_suffix}",
            description=f"Created from {body['src_name']} via Puppyone {spec.kind} → MCP.",
            org_id=org_id,
            created_by=user_id,
            project_limit=project_limit,
            publication_mode="deferred",
            source_fingerprint={
                "kind": "landing-claim",
                "ticket_id": operation_key,
                "tool_kind": spec.kind,
                "source_key": body["md_key"],
                "content_sha256": content_sha256,
            },
            initialize=initialize,
        )
        project = publication.project
        project_id = str(project.id)
        endpoint = endpoint_holder.get("endpoint")
        if endpoint is None:
            # A completed idempotent replay must not create a second endpoint.
            # Plaintext API keys are intentionally not persisted; the response
            # therefore points at the existing endpoint without reissuing one.
            endpoints = await asyncio.to_thread(
                McpEndpointService().list_endpoints,
                project_id,
            )
            endpoint = next(
                (
                    item
                    for item in endpoints
                    if item.get("name") == f"{spec.kind}-mcp"
                    and item.get("path") == spec.scope_path
                ),
                {},
            )
            if not endpoint:
                raise RuntimeError("Completed landing claim is missing its MCP endpoint")
            credential = await asyncio.to_thread(
                AccessCredentialRepository().get_active_by_token,
                api_key,
            )
            if not credential or str(credential.get("access_surface_id")) != str(
                endpoint.get("id")
            ):
                raise AppException(
                    code=ErrorCode.VERSION_CONFLICT,
                    status_code=409,
                    message="The claimed MCP credential was rotated",
                    details={
                        "code": "landing_claim_credential_rotated",
                        "project_id": project_id,
                        "endpoint_id": endpoint.get("id"),
                    },
                )
            endpoint = {
                **endpoint,
                "api_key": api_key,
                "api_key_revealed": True,
            }

        base = (settings.PUBLIC_URL or "").rstrip("/")
        return {
            "project_id": project_id,
            "repo": spec.scope_path,
            "mcp": {
                # External MCP clients connect here; the api_key (Bearer) selects
                # this endpoint.
                "server_url": f"{base}/api/v1/mcp/proxy" if base else "",
                "api_key": endpoint.get("api_key", ""),
                "endpoint_id": endpoint.get("id", ""),
            },
            # Relative — the website prepends its own app origin for the redirect.
            "deep_link": f"/projects/{project_id}",
        }
